#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook

DATE_START = dt.date(2026, 3, 9)
DATE_END = dt.date(2026, 4, 5)
DAY_COL_START = 6
DAY_COL_END = 33

SECTION_ROWS = {
    "OSR": (5, 17),
    "PGY1": (19, 26),
    "PGY2": (28, 37),
    "PGY3": (39, 50),
    "PGY4": (52, 69),
    "ADJ": (71, 78),
}

PGY4_ONLY_CODES = {"D2A", "D2P", "N2"}
NONANCHOR_FRIENDLY = {"D2B", "N2B", "D2", "N2"}
VALID_SHIFT_CODES = {
    # Junior BH/TH
    "D1", "D2", "D2A", "D2B", "D3", "D4", "D4A", "D5",
    "N1", "N2", "N2A", "N2B", "N3", "K2A", "K2B", "TH1A", "T1N", "WC",
    # Senior BH/TH/BK
    "D1A", "D2P", "D3P", "D5P", "D4", "D3", "D5", "D1", "N2", "NF",
    "T1A", "T1P", "T2A", "T2P", "BK9", "BK2", "BKA", "BKN", "DF", "DT", "TF", "TT", "KF",
    # Peds
    "PB8", "PB2", "PB4", "PT8", "PT2", "PT4", "PT9", "PT10", "PT11", "PT12", "PB12",
}
SHIFT_ALIASES = {
    "T110": "T1N",
}


@dataclass
class Resident:
    name: str
    role: str
    anchor_type: Optional[str] = None
    half_assignments: Dict[str, str] = field(default_factory=dict)
    requested_off: Set[dt.date] = field(default_factory=set)
    available_dates: Optional[Set[dt.date]] = None
    amion_id: str = ""


def normalize_name(name: str) -> str:
    s = (name or "").strip().lower()
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def daterange(start: dt.date, end: dt.date) -> List[dt.date]:
    out = []
    d = start
    while d <= end:
        out.append(d)
        d += dt.timedelta(days=1)
    return out


def half_for_date(day: dt.date) -> str:
    return "10A" if day <= dt.date(2026, 3, 22) else "10B"


def parse_requests_text(text: Optional[str]) -> Set[dt.date]:
    out: Set[dt.date] = set()
    if not text:
        return out
    s = str(text)
    for m in re.finditer(r"(\d{1,2})/(\d{1,2})", s):
        mo = int(m.group(1))
        da = int(m.group(2))
        try:
            d = dt.date(2026, mo, da)
        except ValueError:
            continue
        if DATE_START <= d <= DATE_END:
            out.add(d)
    return out


def parse_date_range_text(text: Optional[str]) -> Set[dt.date]:
    out: Set[dt.date] = set()
    if not text:
        return out
    s = str(text).strip()
    m = re.search(r"(\d{1,2})/(\d{1,2})\s*-\s*(\d{1,2})/(\d{1,2})", s)
    if not m:
        return out
    sm, sd, em, ed = map(int, m.groups())
    try:
        start = dt.date(2026, sm, sd)
        end = dt.date(2026, em, ed)
    except ValueError:
        return out
    if end < start:
        return out
    day = start
    while day <= end:
        if DATE_START <= day <= DATE_END:
            out.add(day)
        day += dt.timedelta(days=1)
    return out


def code_site(code: str) -> str:
    c = normalize_shift_code(code)
    if c in {"K2A", "K2B", "TH1A", "T1N", "T1A", "T1P", "T2A", "T2P", "WC", "TT", "TF", "KF"}:
        return "TH"
    if c.startswith("T") and c not in {"DT", "DF"}:
        return "TH"
    if c.startswith("BK"):
        return "BK"
    return "BH"


def code_kind(code: str) -> str:
    c = normalize_shift_code(code)
    if "N" in c and c not in {"D1", "D2A", "D2P", "D3", "D4", "D5", "D5P", "DF", "DT"}:
        return "night"
    if c.startswith("N") or c.endswith("N"):
        return "night"
    if c in {"D2A", "D4A", "K2A", "TH1A", "T1A", "T2A", "DT", "TF", "PB8", "PT8", "PT9", "PT10", "PT11", "BK9", "BKA"}:
        return "day"
    return "swing"


def normalize_shift_code(raw: str) -> str:
    c = re.sub(r"[^A-Z0-9]", "", str(raw).strip().upper())
    return SHIFT_ALIASES.get(c, c)


def is_valid_shift_code(raw: str) -> bool:
    return normalize_shift_code(raw) in VALID_SHIFT_CODES


def role_can_work(res: Resident, day: dt.date, code: str, target_role: str) -> bool:
    hb = half_for_date(day)
    assignment = res.half_assignments.get(hb, "")
    c = normalize_shift_code(code)
    site = code_site(c)

    if target_role in {"OSR", "PGY1", "PGY2", "PGY3", "PGY4"} and res.role != target_role:
        return False
    if res.available_dates is not None and day not in res.available_dates:
        return False

    if res.role == "OSR":
        if site != "BH":
            return False
        if res.anchor_type == "nonanchor" and c not in NONANCHOR_FRIENDLY:
            return False

    if res.role == "PGY1":
        if assignment not in {"BH", "TH"}:
            return False
        if site == "BH" and assignment != "BH":
            return False
        if site == "TH" and assignment != "TH":
            return False
        if site == "BK":
            return False

    if res.role == "PGY2":
        if assignment != "BH/TH":
            return False
        if site == "BK":
            return False

    if res.role in {"PGY3", "PGY4"}:
        if "ED" not in assignment.upper():
            return False

    if c in PGY4_ONLY_CODES and res.role != "PGY4":
        return False

    return True


def max_shifts(res: Resident) -> int:
    if res.role == "PGY4":
        return 14
    if res.role == "OSR":
        return 20
    return 18


def assign_slots(
    residents: Dict[str, Resident],
    slots_by_role: Dict[str, Dict[dt.date, List[str]]],
) -> Tuple[Dict[str, Dict[dt.date, str]], Dict[str, Dict[str, int]], int]:
    by_res_day: Dict[str, Dict[dt.date, str]] = defaultdict(dict)
    counts_half: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    last_kind: Dict[str, Tuple[dt.date, str]] = {}
    conflicts = 0

    ordered_roles = ["OSR", "PGY1", "PGY2", "PGY3", "PGY4", "ADJ"]

    for role in ordered_roles:
        for day in sorted(slots_by_role.get(role, {}).keys()):
            codes = slots_by_role[role][day]
            codes = sorted(codes, key=lambda x: (0 if code_kind(x) == "night" else 1, x))

            for code in codes:
                candidates: List[Tuple[float, str]] = []
                for key, res in residents.items():
                    if day in by_res_day[key]:
                        continue

                    target_role = role
                    if role == "ADJ":
                        target_role = res.role

                    if not role_can_work(res, day, code, target_role):
                        continue

                    hb = half_for_date(day)
                    if counts_half[key][hb] >= max_shifts(res):
                        continue

                    lk = last_kind.get(key)
                    if lk:
                        last_day, last_k = lk
                        if last_day == day - dt.timedelta(days=1) and last_k == "night" and code_kind(code) in {"day", "swing"}:
                            continue

                    score = 0.0
                    if day in res.requested_off:
                        score += 100.0

                    score += counts_half[key][hb] * 2.5

                    if res.role == "PGY2" and code_site(code) == "TH":
                        score += 2.0

                    if res.role == "OSR" and res.anchor_type == "nonanchor":
                        if (code or "").upper() in NONANCHOR_FRIENDLY:
                            score -= 1.0

                    candidates.append((score, key))

                if not candidates:
                    continue

                candidates.sort(key=lambda x: (x[0], x[1]))
                chosen = candidates[0][1]
                by_res_day[chosen][day] = code
                hb = half_for_date(day)
                counts_half[chosen][hb] += 1
                last_kind[chosen] = (day, code_kind(code))
                if day in residents[chosen].requested_off:
                    conflicts += 1

    return by_res_day, counts_half, conflicts


def make_amion(name: str) -> str:
    parts = [p for p in re.split(r"\s+", re.sub(r"\([^)]*\)", "", name.strip())) if p]
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0][:3].upper()
    return (parts[0][0] + parts[-1][:2]).upper()


def load_template_amion_map(ws) -> Dict[str, str]:
    m = {}
    for r in range(5, 79):
        n = ws.cell(r, 1).value
        a = ws.cell(r, 2).value
        if n and a:
            m[normalize_name(str(n))] = str(a)
    return m


def load_roster(roster_path: Path) -> Dict[str, Resident]:
    wb = load_workbook(roster_path, data_only=True)
    residents: Dict[str, Resident] = {}

    def colidx(ws, label: str) -> int:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        return headers.index(label) + 1

    for role in ["PGY1", "PGY2", "PGY3", "PGY4"]:
        ws = wb[role]
        c9b = colidx(ws, "9B")
        c10a = colidx(ws, "10A")
        c10b = colidx(ws, "10B")
        c11a = colidx(ws, "11A")

        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name:
                continue
            name = str(name).strip()
            a10 = str(ws.cell(r, c10a).value or "")
            b10 = str(ws.cell(r, c10b).value or "")

            include = False
            if role == "PGY1":
                include = a10 in {"BH", "TH"} or b10 in {"BH", "TH"}
            elif role == "PGY2":
                include = a10 == "BH/TH" or b10 == "BH/TH"
            else:
                include = ("ED" in a10.upper()) or ("ED" in b10.upper())
            if not include:
                continue

            res = Resident(
                name=name,
                role=role,
                half_assignments={
                    "9B": str(ws.cell(r, c9b).value or ""),
                    "10A": a10,
                    "10B": b10,
                    "11A": str(ws.cell(r, c11a).value or ""),
                },
            )
            residents[normalize_name(name)] = res

    return residents


def load_osrs(osr_path: Path) -> Dict[str, Resident]:
    wb = load_workbook(osr_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: Dict[str, Resident] = {}
    for r in range(2, ws.max_row + 1):
        raw = ws.cell(r, 1).value
        if not raw:
            continue
        raw = str(raw).strip()
        anchor = "nonanchor" if re.search(r"ophthal|ophth|psych|omfs", raw, flags=re.I) else "anchor"
        availability = parse_date_range_text(ws.cell(r, 3).value)
        res = Resident(
            name=raw,
            role="OSR",
            anchor_type=anchor,
            half_assignments={"10A": "BH", "10B": "BH"},
            available_dates=availability if availability else set(daterange(DATE_START, DATE_END)),
        )
        amion = ws.cell(r, 2).value
        if amion not in (None, ""):
            res.amion_id = str(amion).strip()
        out[normalize_name(raw)] = res
    return out


def resolve_existing_path(*candidates: str) -> str:
    for c in candidates:
        if Path(c).exists():
            return c
    return candidates[0]


def load_requests(req_path: Path) -> Dict[str, Set[dt.date]]:
    wb = load_workbook(req_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    reqs: Dict[str, Set[dt.date]] = defaultdict(set)
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 3).value
        dates = ws.cell(r, 7).value
        if not name:
            continue
        reqs[normalize_name(str(name))].update(parse_requests_text(dates))
    return reqs


def extract_template_slots(ws, template_ws) -> Dict[str, Dict[dt.date, List[str]]]:
    # Read from pristine template sheet so slot structure stays tied to the model workbook.
    days = daterange(DATE_START, DATE_END)
    slots: Dict[str, Dict[dt.date, List[str]]] = defaultdict(lambda: defaultdict(list))
    for role, (r1, r2) in SECTION_ROWS.items():
        for row in range(r1, r2 + 1):
            for ci, day in enumerate(days, start=DAY_COL_START):
                val = template_ws.cell(row, ci).value
                if val not in (None, ""):
                    slots[role][day].append(str(val).strip())
    return slots


def clear_section_cells(ws):
    for _, (r1, r2) in SECTION_ROWS.items():
        for row in range(r1, r2 + 1):
            for col in range(1, DAY_COL_END + 1):
                cell = ws.cell(row, col)
                if cell.value not in (None, ""):
                    cell.value = None


def set_headers(ws):
    days = daterange(DATE_START, DATE_END)
    ws.cell(2, 1).value = "Block 10"
    ws.cell(3, 1).value = "03/09/26-04/05/26"
    dow_map = {0: "M", 1: "T", 2: "W", 3: "Th", 4: "F", 5: "Sa", 6: "Su"}
    for i, day in enumerate(days, start=DAY_COL_START):
        ws.cell(2, i).value = dow_map[day.weekday()]
        ws.cell(3, i).value = float(day.day)


def resident_lists_by_role(residents: Dict[str, Resident]) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = defaultdict(list)
    for key, res in residents.items():
        out[res.role].append(key)
    for k in out:
        out[k] = sorted(out[k], key=lambda x: residents[x].name)
    return out


def assign_rows(residents: Dict[str, Resident]) -> Tuple[Dict[int, str], Dict[str, int]]:
    row_to_res: Dict[int, str] = {}
    res_to_row: Dict[str, int] = {}

    by_role = resident_lists_by_role(residents)
    overflow: List[str] = []

    for role in ["OSR", "PGY1", "PGY2", "PGY3", "PGY4"]:
        rows = list(range(SECTION_ROWS[role][0], SECTION_ROWS[role][1] + 1))
        keys = by_role.get(role, [])
        for row, key in zip(rows, keys):
            row_to_res[row] = key
            res_to_row[key] = row
        if len(keys) > len(rows):
            overflow.extend(keys[len(rows) :])

    adj_rows = list(range(SECTION_ROWS["ADJ"][0], SECTION_ROWS["ADJ"][1] + 1))
    for row, key in zip(adj_rows, overflow):
        row_to_res[row] = key
        res_to_row[key] = row

    return row_to_res, res_to_row


def row_pattern(template_ws, row: int) -> List[Tuple[dt.date, str]]:
    out: List[Tuple[dt.date, str]] = []
    days = daterange(DATE_START, DATE_END)
    for ci, day in enumerate(days, start=DAY_COL_START):
        val = template_ws.cell(row, ci).value
        if val not in (None, "") and is_valid_shift_code(str(val)):
            out.append((day, str(val).strip()))
    return out


def assign_residents_to_template_rows(
    residents: Dict[str, Resident], template_ws
) -> Tuple[Dict[int, str], Dict[str, int], int]:
    row_to_res: Dict[int, str] = {}
    used: Set[str] = set()
    request_conflicts = 0

    def pick_for_row(role: str, row: int, pool: List[str]) -> Optional[str]:
        pattern = row_pattern(template_ws, row)
        if not pattern:
            return None
        candidates: List[Tuple[float, str, int]] = []
        for key in pool:
            if key in used:
                continue
            res = residents[key]
            if role != "ADJ" and res.role != role:
                continue

            hard_bad = 0
            bad = 0
            req = 0
            for day, code in pattern:
                if res.role == "OSR" and res.available_dates is not None and day not in res.available_dates:
                    hard_bad += 1
                    continue
                target = role if role != "ADJ" else res.role
                if not role_can_work(res, day, code, target):
                    bad += 1
                    continue
                if day in res.requested_off:
                    req += 1
            if hard_bad > 0:
                continue
            score = bad * 50 + req * 100 + len(pattern)
            candidates.append((score, key, req))
        if not candidates:
            return None
        candidates.sort(key=lambda x: (x[0], x[1]))
        best = candidates[0]
        return best[1]

    by_role: Dict[str, List[str]] = defaultdict(list)
    for key, res in residents.items():
        by_role[res.role].append(key)
    for role in by_role:
        by_role[role].sort(key=lambda k: residents[k].name)

    # Fill primary sections first.
    for role in ["OSR", "PGY1", "PGY2", "PGY3", "PGY4"]:
        r1, r2 = SECTION_ROWS[role]
        pool = by_role.get(role, [])
        for row in range(r1, r2 + 1):
            chosen = pick_for_row(role, row, pool)
            if not chosen:
                continue
            row_to_res[row] = chosen
            used.add(chosen)
            # Count request conflicts for this matched row pattern.
            req = sum(1 for day, _ in row_pattern(template_ws, row) if day in residents[chosen].requested_off)
            request_conflicts += req

    # Use adjunct rows as overflow for any remaining resident.
    overflow_pool = [k for k in sorted(residents.keys(), key=lambda k: residents[k].name) if k not in used]
    r1, r2 = SECTION_ROWS["ADJ"]
    for row in range(r1, r2 + 1):
        chosen = pick_for_row("ADJ", row, overflow_pool)
        if not chosen:
            continue
        row_to_res[row] = chosen
        used.add(chosen)
        req = sum(1 for day, _ in row_pattern(template_ws, row) if day in residents[chosen].requested_off)
        request_conflicts += req

    stats: Dict[str, int] = defaultdict(int)
    for row, key in row_to_res.items():
        stats[f"{residents[key].role}_rows"] += 1
        stats["assigned_rows"] += 1

    return row_to_res, dict(stats), request_conflicts


def fill_bhth_sheet(
    ws,
    template_ws,
    residents: Dict[str, Resident],
    row_to_res: Dict[int, str],
    amion_map: Dict[str, str],
):
    def write_cell(row: int, col: int, value):
        cell = ws.cell(row, col)
        cell._style = template_ws.cell(row, col)._style
        cell.value = value

    set_headers(ws)

    # Rewrite section header rows exactly where they exist in the template layout.
    write_cell(4, 1, "OSR")
    write_cell(4, 2, "Amion ID")
    write_cell(4, 3, "Before")
    write_cell(4, 4, "Block")
    write_cell(4, 5, "After")

    write_cell(18, 1, "PGY-1")
    write_cell(18, 2, "Amion ID")
    write_cell(18, 3, "Before")
    write_cell(18, 4, "Block")
    write_cell(18, 5, "After")

    write_cell(27, 1, "PGY-2")
    write_cell(27, 2, "Amion ID")
    write_cell(27, 3, "Before")
    write_cell(27, 4, "Block")
    write_cell(27, 5, "After")

    write_cell(38, 1, "PGY-3")
    write_cell(38, 2, "Amion ID")
    write_cell(38, 3, "Before")
    write_cell(38, 4, "Block")
    write_cell(38, 5, "After")

    write_cell(51, 1, "PGY-4")
    write_cell(51, 2, "Amion ID")
    write_cell(51, 3, "Before")
    write_cell(51, 4, "Block")
    write_cell(51, 5, "After")

    write_cell(70, 1, "Adjunct Residents")
    write_cell(70, 2, "Amion ID")
    write_cell(70, 3, "Before")
    write_cell(70, 4, "Block")
    write_cell(70, 5, "After")

    for row in range(5, 79):
        key = row_to_res.get(row)
        # Always clear day-grid values; preserve formatting, remove template artifacts.
        for ci in range(DAY_COL_START, DAY_COL_END + 1):
            cell = ws.cell(row, ci)
            if cell.value not in (None, ""):
                cell.value = None

        if not key:
            # Keep formatting but clear values for unused rows.
            for col in range(1, DAY_COL_END + 1):
                cell = ws.cell(row, col)
                if cell.value not in (None, ""):
                    cell.value = None
            continue
        res = residents[key]

        write_cell(row, 1, res.name)
        nid = res.amion_id or amion_map.get(key, "") or make_amion(res.name)
        write_cell(row, 2, nid)
        res.amion_id = nid

        before = res.half_assignments.get("9B", "")
        b10a = res.half_assignments.get("10A", "")
        b10b = res.half_assignments.get("10B", "")
        after = res.half_assignments.get("11A", "")
        write_cell(row, 3, before)
        write_cell(row, 4, f"{b10a} / {b10b}" if (b10a or b10b) else "")
        write_cell(row, 5, after)

        # Preserve template cell positions/colors by writing only where the template row has shifts.
        for ci in range(DAY_COL_START, DAY_COL_END + 1):
            shift = template_ws.cell(row, ci).value
            if shift not in (None, "") and is_valid_shift_code(str(shift)):
                write_cell(row, ci, shift)


def update_osr_sheet(wb, residents: Dict[str, Resident]):
    if "OSRs" not in wb.sheetnames:
        return
    ws = wb["OSRs"]
    ws.cell(1, 1).value = "Off service rotator"
    ws.cell(1, 2).value = "Amion ID"

    osrs = [r for r in residents.values() if r.role == "OSR"]
    osrs.sort(key=lambda x: x.name)
    for i, r in enumerate(osrs, start=2):
        ws.cell(i, 1).value = r.name
        ws.cell(i, 2).value = r.amion_id or make_amion(r.name)


def schedule_summary(
    residents: Dict[str, Resident],
    template_ws,
    row_to_res: Dict[int, str],
    row_stats: Dict[str, int],
    conflicts: int,
) -> Dict[str, int]:
    total_assignments = sum(len(row_pattern(template_ws, row)) for row in row_to_res)
    out = {"conflicts": conflicts, "assignments": total_assignments, "assigned_rows": row_stats.get("assigned_rows", 0)}
    for role in ["OSR", "PGY1", "PGY2", "PGY3", "PGY4"]:
        keys = [k for k, r in residents.items() if r.role == role]
        out[f"{role}_residents"] = len(keys)
        out[f"{role}_rows"] = row_stats.get(f"{role}_rows", 0)
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate Block 10 schedule using Block 9 formatting template.")
    parser.add_argument("--template", default=resolve_existing_path("Example/block 9 schedule.xlsx", "Example/block 9 output.xlsx"))
    parser.add_argument("--roster", default="25-26 schedule.xlsx")
    parser.add_argument("--osr", default=resolve_existing_path("Test/block 10 OSRs.xlsx", "Test/block 10 data.xlsx"))
    parser.add_argument("--requests", default=resolve_existing_path("Test/block 10 scheduling requests.xlsx", "Test/SAMN Block 10 Request Form (Responses)  (claude test).xlsx"))
    parser.add_argument("--output", default="Test/block 10 generated schedule.xlsx")
    args = parser.parse_args()

    wb = load_workbook(args.template)
    template_wb = load_workbook(args.template, data_only=False)
    ws = wb["BHTH"]
    template_ws = template_wb["BHTH"]

    amion_existing = load_template_amion_map(template_ws)

    residents = load_roster(Path(args.roster))
    residents.update(load_osrs(Path(args.osr)))

    reqs = load_requests(Path(args.requests))
    for key, days in reqs.items():
        if key in residents:
            residents[key].requested_off.update(days)

    row_to_res, row_stats, conflicts = assign_residents_to_template_rows(residents, template_ws)
    fill_bhth_sheet(ws, template_ws, residents, row_to_res, amion_existing)
    update_osr_sheet(wb, residents)
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    stats = schedule_summary(residents, template_ws, row_to_res, row_stats, conflicts)

    out = Path(args.output)
    wb.save(out)

    print(f"Wrote {out}")
    print(f"Residents: {len(residents)}")
    print(f"Assignments: {stats['assignments']}")
    print(f"Request conflicts: {stats['conflicts']}")
    for role in ["OSR", "PGY1", "PGY2", "PGY3", "PGY4"]:
        print(f"{role}: residents={stats[f'{role}_residents']}, rows_used={stats[f'{role}_rows']}")


if __name__ == "__main__":
    main()
